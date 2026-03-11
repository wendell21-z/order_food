from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame

work_shop_to_pos = {
    "总装车间": "YP_0031500003",
    "焊装车间": "YP_0031500002",
}


class Bill:
    def __init__(self, amount: float, balance: float, time, pos: str):
        # 消费金额
        self.amount = float(amount)
        # 支付后余额
        self.balance = float(balance)
        # 发生时间
        self.time = time
        # pos
        self.pos = pos


class ErrorBill(Bill):
    """
    没有订单的支付信息
    """

    def __init__(self, work_no, name, bill, pos):
        super().__init__(bill.amount, bill.balance, bill.time, pos)
        self.work_no: int = work_no
        self.name = name
        self.bill = bill


class Order:
    def __init__(self, work_no, name, workshop, workshop_section, take_out_point):
        # 工号
        self.work_no: int = work_no
        # 姓名
        self.name = name
        # 取餐车间
        self.workshop = workshop
        # 车间-工段
        self.workshop_section = workshop_section
        # 取餐点
        self.take_out_point = take_out_point
        # dishes
        self.dishes: List[Dish] = []
        # 支付状态。 值小于0表示支付金额不足，值大于0表示支付金额超出
        self.pay_status = ''
        # 总花费
        self.total_pay: float = 0.0
        # 菜品总价格
        self.total_price: float = 0.0
        # 金额差值
        self.amount_diff: float = 0.0
        # 支付信息
        self.bills: List[Bill] = []
        # 余额
        self.balance = 0

    def calc_money(self):
        self.calc_total_pay()
        self.calc_total_price()
        self.calc_amount_diff()
        self.calc_balance()

    def calc_total_pay(self):
        self.total_pay = sum([(-bill.amount) for bill in self.bills])

    def calc_total_price(self):
        self.total_price = sum([dish.num * dish.price for dish in self.dishes])

    def calc_amount_diff(self):
        self.amount_diff = (self.total_pay - self.total_price)
        if self.amount_diff < 0:
            self.pay_status = '支付金额不足'
        elif self.amount_diff > 0:
            self.pay_status = '支付金额超出'
        else:
            self.pay_status = '正常'

    def calc_balance(self):
        """
        bill里time最大的元素的balance
        :return:
        """
        if len(self.bills) > 0:
            self.balance = float(max(self.bills, key=lambda x: x.time).balance)
        else:
            self.balance = 0.0


class Dish:
    def __init__(self, name, num: int, price: float):
        # 菜名
        self.name: str = name
        # 数量
        self.num = int(num)
        # 单价
        self.price = float(price)


# pase excel to order
def create_order(file_path) -> Dict[int, Order]:
    """
    从下单excel 创建 订单对象
    :param file_path: excel路径
    :return: {work_no: Order}
    """

    r: Dict[int, Order] = {}  # work_no -> Order

    df = read_excel(file_path, 2)

    for index, row in df.iterrows():

        work_no = int(row[('员工工号', '员工工号')])

        name = row[('员工姓名', '员工姓名')]
        workshop = row[('取餐车间', '取餐车间')]
        workshop_section = row[('车间-工段', '车间-工段')]
        if workshop == '总装车间':
            take_out_point = workshop_section
        elif workshop == '涂装车间':
            take_out_point = f'{row[('取餐点', '取餐点')]}-{workshop_section}'
        else:
            take_out_point = row[('取餐点', '取餐点')]

        if work_no not in r:
            r[work_no] = Order(work_no, name, workshop, workshop_section, take_out_point)

        order = r[work_no]
        order.dishes.append(Dish(row[('加餐明细', '菜单')], row[('加餐明细', '购买数量')], row[('加餐明细', '单价')]))
    return r


def add_bill_info(order_map: Dict[int, Order], file_path) -> List[ErrorBill]:
    """
    添加订单信息。
    :param order_map: 订单信息map
    :param file_path: 支付信息excel
    :return: 错误支付。有支付信息，但没有订单
    """
    error_bills = []

    df = read_excel(file_path)

    # 筛选
    # 1. pos机 = YP_0031500002
    # 2. 业务种类 = 消费
    if '业务种类' in df.columns:
        df = df[df['业务种类'] == '消费']

    for index, row in df.iterrows():
        # 支付机号
        bill_pos = row.get('POS机号') or row.get('设备编号', '')
        # 工号
        work_no_str = row['员工号']
        # 如果工号不能被转为int，说明不是公司员工，跳过
        if not work_no_str.isdigit():
            continue
        work_no = int(work_no_str)
        balance = row.get('发生后库余额', 0.0)
        pay_time = row.get('发生时间') or row.get('消费时间', '')
        expenditure = row.get('发生额') or row.get('消费金额', 0.0)
        bill = Bill(expenditure, balance, pay_time, bill_pos)

        # 处理有支付信息，没有订单信息的情况
        if work_no not in order_map:
            if bill_pos in work_shop_to_pos.values():
                name = row.get('姓名') or row.get('员工姓名', '')
                error_bills.append(ErrorBill(work_no, name, bill, bill_pos))
            continue

        # 只有对应卡机的支付信息可以添加
        order = order_map[work_no]
        # TODO 按卡机过滤，暂时注释
        # work_shop = order.workshop
        # order_pos = work_shop_to_pos[work_shop]
        # if bill_pos != order_pos:
        #     continue
        order.bills.append(bill)

    return error_bills


def read_excel(file_path, n_header: int = 1) -> DataFrame:
    """

    :param file_path:
    :param n_header: 前n行数据是header
    :return:
    """
    # 使用 openpyxl 加载 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb.active
    df = DataFrame(sheet.values)

    # 获取所有合并单元格的范围
    merged_cells = sheet.merged_cells

    # 遍历所有合并单元格区域，并填充 NaN 区域
    for merged_cell in merged_cells:
        min_col, min_row, max_col, max_row = merged_cell.bounds
        merged_value = sheet.cell(row=min_row, column=min_col).value

        # if merged_value is None:
        #     continue

        for row in range(min_row - 1, max_row):  # pandas 使用 1-based 索引
            for col in range(min_col - 1, max_col):  # pandas 使用 1-based 索引
                df.iloc[row, col] = merged_value

    if n_header < 1:
        return df
    if n_header < 2:
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    else:
        # 将前两行作为多级列索引
        df.columns = pd.MultiIndex.from_arrays(df.iloc[:n_header].values)
        # 保留前两行之后的数据，并重新排序
        df = df.iloc[n_header:].reset_index(drop=True)

    return df


if __name__ == '__main__':
    g_order_map = create_order('data/员工加餐购买小程序_20251228101936.xlsx')
    error_bill_map = add_bill_info(g_order_map, 'data/sigma_export (96).xlsx')
    for g_order in g_order_map.values():
        g_order.calc_amount_diff()
        g_order.calc_balance()
    print(g_order_map)
